import openpyxl
import xlwt
from openpyxl.utils import get_column_letter

# 定义原始表格文件路径列表
file_paths = [
    r'/Users/wacai/Desktop/d2/mircosoft.xlsx',
   
    # 添加更多文件路径...
]

# 定义目标表格A的文件路径
target_path = '/Users/wacai/Desktop/d2/1.xlsx'

# 定义目标表格A的列名和对应的列字母
target_columns = {
    "结算日期": "A",
    "产品": "B",
    "保底": "D",
    "有效订单数": "F"
}

# 加载目标表格A
target_workbook = openpyxl.load_workbook(target_path)

# 获取目标表格A的工作表（假设在第一个工作表）
target_sheet = target_workbook.active
source_sheet1=xlwt.Workbook()

# for column_name, column_letter in target_columns.items():
#     print(column_name)
#     print(column_letter)
#     print(target_sheet.max_row+1)
    # print

# 逐个处理每个文件
for file_path in file_paths:
    # 打开当前文件
    workbook = openpyxl.load_workbook(file_path)
    
    # 获取当前工作表（假设在第一个工作表）
    source_sheet = workbook.active
    times=1
    # 复制每一行的数据到目标表格A的指定列中
    for row in source_sheet.iter_rows(values_only=True):
        

        # print(row)
        if times==0:
            pass
        # print(column_name,column_letter)
        else:  
            # print(target_sheet['A1'].value)

            target_sheet[f'A{times}'].value=row[1]
            target_sheet.cell(row=times,column=2,value=row[2])
            
        times+=1
            # source_sheet.write()
            # target_sheet[column_letter + str(target_sheet.max_row + 1)].value = row[source_sheet[column_name].column - 1]

# 保存目标表格A
target_workbook.save("merged_table_A.xlsx")