from openpyxl import load_workbook,Workbook

# 加载 Excel 文件
workbook = load_workbook('/Users/wacai/Desktop/d2/output(1).xlsx')


# 选择工作表
sheet = workbook.active

# 获取某一列的数值列表
column_values = [cell.value.strip() for cell in sheet['D']]  # 假设列为 A
# column_values=column_values;
# 使用集合去重
unique_values = list(set(column_values))

# 创建新的工作簿对象
new_workbook = Workbook()

# 获取默认的工作表
new_sheet = new_workbook.active

# 将去重后的数据写入新的工作表
for i, value in enumerate(unique_values):
    
    new_sheet.cell(row=i+1, column=1, value=value)  # 写入第一列

# 保存工作簿到文件
new_workbook.save('ai_respone.xlsx')
